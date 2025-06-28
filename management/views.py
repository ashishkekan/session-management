from datetime import datetime

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from management.forms import (
    CustomPasswordChangeForm,
    DepartmentForm,
    ExternalTopicForm,
    SessionTopicForm,
    SessionUploadForm,
    UserCreationForm,
    UserEditForm,
)
from management.models import (
    PLACE_CHOICES,
    STATUSES,
    Department,
    ExternalTopic,
    RecentActivity,
    SessionTopic,
)
from management.utils import log_activity


@login_required
def create_topic(request):
    """
    Create a new session topic.

    - Staff users can create sessions for others and notify normal users.
    - Normal users creating a session notify staff users.

    Returns:
        HttpResponse: Rendered page with the topic form.
    """
    if request.method == "POST":
        form = SessionTopicForm(request.POST or None, user=request.user)
        if form.is_valid():
            session = form.save()
            if request.user.is_staff:
                normal_users = User.objects.filter(is_staff=False)
                log_activity(
                    request.user,
                    f"Admin created a new session: '{session.topic}'.",
                    target_users=normal_users,
                )
            else:
                # Normal user action, log_activity will notify admins automatically
                log_activity(request.user, f"Created a new session: '{session.topic}'.")
            return redirect("session_list")
    else:
        form = SessionTopicForm(user=request.user)
    return render(request, "session/create_topic.html", {"form": form})


def home(request):
    """
    Display the home page.

    - For staff: shows user/session stats and upcoming sessions.
    - For regular users: shows user's upcoming sessions.

    Returns:
        HttpResponse: Rendered home page with context.
    """
    user = request.user
    latest_topics = ExternalTopic.objects.order_by("-created_at")

    context = {
        "learning_topics": latest_topics,
    }
    top_sessions = (
        SessionTopic.objects.filter(
            date__gt=now(),
        )
        .exclude(status__in=["Completed", "Cancelled"])
        .select_related("conducted_by")
        .order_by("date")[:3]
    )

    top_sessions = (
        SessionTopic.objects.filter(date__gt=now())
        .exclude(status__in=["Completed", "Cancelled"])
        .select_related("conducted_by")
        .order_by("date")[:3]
    )

    if user.is_staff:
        context.update(
            {
                "is_admin": True,
                "total_users": User.objects.count(),
                "total_sessions": SessionTopic.objects.count(),
                "all_sessions": SessionTopic.objects.order_by("-date"),
                "top_sessions": top_sessions,
                "completed": SessionTopic.objects.filter(status="Completed").order_by(
                    "date"
                )[:3],
                "pending": SessionTopic.objects.filter(status="Pending").order_by(
                    "date"
                )[:3],
                "cancelled": SessionTopic.objects.filter(status="Cancelled").order_by(
                    "date"
                )[:3],
            }
        )
    elif user.is_authenticated:
        sessions = SessionTopic.objects.filter(conducted_by=user)
        upcoming_sessions = sessions.filter(status="Pending", date__gte=now()).order_by(
            "date"
        )
        context.update(
            {
                "is_admin": False,
                "total_sessions": sessions.count(),
                "upcoming_sessions": upcoming_sessions,
                "top_sessions": top_sessions,
            }
        )

    return render(request, "session/home.html", context)


def user_login(request):
    """
    Handle user login.

    Authenticates credentials and redirects to home upon success.

    Returns:
        HttpResponse: Login form or redirect on success.
    """
    if request.method == "POST":
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            log_activity(request.user, "Logged in successfully.")
            return redirect("home")
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, "session/login.html", {"form": form})


def user_logout(request):
    """
    Log the user out and redirect to login.

    Returns:
        HttpResponseRedirect: Redirect to login page.
    """
    if request.user.is_authenticated:
        log_activity(request.user, "Logged out.")
    logout(request)
    return redirect("login")


def is_admin(user):
    """Check if the user is an admin."""
    return user.is_staff


@login_required
@user_passes_test(is_admin)
def add_user(request):
    """
    Admins can add a new user to the system.

    Returns:
        HttpResponse: User creation form or redirect on success.
    """
    if request.method == "POST":
        form = UserCreationForm(request.POST or None)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            log_activity(
                request.user,
                f"Admin added new user '{user.username}'.",
                target_users=User.objects.filter(is_staff=False),
            )
            messages.success(request, f"User '{user.username}' created successfully.")
            return redirect("home")
    else:
        form = UserCreationForm()
    return render(request, "session/add_user.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def edit_user(request, user_id):
    """
    Admins can edit a user’s profile.

    Args:
        user_id (int): ID of the user to edit.

    Returns:
        HttpResponse: Edit form or redirect on success.
    """
    user = User.objects.get(id=user_id)
    if request.method == "POST":
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            log_activity(request.user, "Admin edited your profile.", edited_user=user)
            messages.success(request, "User updated successfully.")
            return redirect("user_list")
    else:
        form = UserEditForm(instance=user)
    return render(request, "session/edit_user.html", {"form": form, "user_obj": user})


@login_required
@user_passes_test(is_admin)
def delete_user(request, user_id):
    """
    Delete a user from the system (only by admin).

    Args:
        user_id (int): ID of the user to delete.

    Returns:
        HttpResponseRedirect: Redirect to user list.
    """
    user = get_object_or_404(User, id=user_id)
    if user.is_staff:
        messages.error(request, "You cannot delete a superuser.")
        return redirect("user_list")
    user.delete()
    log_activity(
        request.user,
        f"Admin deleted user with ID {user_id}.",
        target_users=User.objects.filter(is_staff=False),
    )
    messages.success(request, "User deleted successfully.")
    return redirect("user_list")


@login_required
def my_profile(request):
    """
    View and update the logged-in user's profile.

    Returns:
        HttpResponse: Profile page.
    """
    employee = request.user
    if request.method == "POST":
        form = UserEditForm(request.POST, instance=employee)
        if not form.is_valid():
            messages.error(request, "There was an error updating your profile.")
        form.save()
        log_activity(employee, "Updated their profile.")
        messages.success(request, "Profile updated successfully.")
        return redirect("my_profile")
    form = UserEditForm(instance=employee)
    return render(request, "session/my_profile.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def user_list(request):
    """
    Display a paginated list of registered users for admin users.

    Allows filtering users by department via a dropdown selection.
    If a department is selected (via GET parameter), only users associated
    with that department will be displayed.

    Context passed to template:
        - users: Paginated queryset of users.
        - departments: All departments for the dropdown filter.
        - selected_department: Currently selected department ID (if any).

    Template:
        session/user_list.html

    Returns:
        HttpResponse: Rendered template showing the list of users.
    """
    department_id = request.GET.get("department")
    departments = Department.objects.all()

    users = User.objects.all().order_by("username").select_related("userprofile")

    if department_id:
        users = users.filter(userprofile__department_id=department_id)

    paginator = Paginator(users, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "session/user_list.html",
        {
            "users": page_obj,
            "departments": departments,
            "selected_department": department_id,
        },
    )


@login_required
def all_sessions_view(request):
    """
    View all sessions.

    - Admins see all sessions, filtered by search query if provided.
    - Regular users see only their own sessions, filtered by search query if provided.

    Args:
        request: HTTP request object.

    Returns:
        HttpResponse: Rendered list of sessions with pagination.
    """
    # Get the search query from the GET parameters
    search_query = request.GET.get("q", "").strip()

    if request.user.is_staff:
        sessions = SessionTopic.objects.select_related("conducted_by").order_by("date")
    else:
        sessions = (
            SessionTopic.objects.select_related("conducted_by")
            .filter(conducted_by=request.user)
            .order_by("date")
        )

    # Apply search filter if query is provided
    if search_query:
        sessions = sessions.filter(
            Q(topic__icontains=search_query)
            | Q(status__icontains=search_query)
            | Q(place__icontains=search_query)
            | Q(conducted_by__first_name__icontains=search_query)
            | Q(conducted_by__last_name__icontains=search_query)
            | Q(cancelled_reason__icontains=search_query)
        )

    # Paginate the results
    paginator = Paginator(sessions, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "session/all_sessions.html", {"sessions": page_obj})


@login_required
def edit_session_view(request, session_id):
    """
    Edit an existing session.

    Args:
        session_id (int): ID of the session.

    Returns:
        HttpResponse: Form or redirect.
    """
    session = get_object_or_404(SessionTopic, id=session_id)
    if request.method == "POST":
        form = SessionTopicForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            if request.user.is_staff:
                log_activity(
                    request.user,
                    f"Admin updated session: '{session.topic}'.",
                    target_users=User.objects.filter(is_staff=False),
                )
            else:
                log_activity(request.user, f"Updated session: '{session.topic}'.")
            messages.success(request, "Session updated successfully.")
            return redirect("session_list")
    else:
        form = SessionTopicForm(instance=session)
    return render(
        request, "session/edit_session.html", {"form": form, "session": session}
    )


@login_required
def delete_session_view(request, session_id):
    """
    Delete a session.

    Args:
        session_id (int): ID of the session.

    Returns:
        HttpResponseRedirect: Redirect to session list.
    """
    session = get_object_or_404(SessionTopic, id=session_id)
    if request.user.is_staff:
        log_activity(
            request.user,
            f"Admin deleted session: '{session.topic}'.",
            target_users=User.objects.filter(is_staff=False),
        )
    else:
        log_activity(request.user, f"Deleted session: '{session.topic}'.")
    session.delete()
    messages.success(request, "Session deleted successfully.")
    return redirect("session_list")


@login_required
def change_password(request):
    """
    Change the current user's password.

    Returns:
        HttpResponse: Password change form or redirect.
    """
    if request.method == "POST":
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            log_activity(request.user, "Changed their password.")
            messages.success(request, "Your password was successfully updated!")
            return redirect("my_profile")
    else:
        form = CustomPasswordChangeForm(user=request.user)
    return render(request, "session/change_password.html", {"form": form})


@login_required
def create_external_topic(request):
    """
    Create a new external learning topic.

    Returns:
        HttpResponse: Form or redirect.
    """
    if request.method == "POST":
        form = ExternalTopicForm(request.POST)
        if form.is_valid():
            topic = form.save()
            if request.user.is_staff:
                log_activity(
                    request.user,
                    f"Admin added new learning topic: '{topic.coming_soon}'.",
                    target_users=User.objects.filter(is_staff=False),
                )
            else:
                log_activity(
                    request.user, f"Added new learning topic: '{topic.coming_soon}'."
                )
            messages.success(request, "New topic created successfully.")
            return redirect("learning-view")
    else:
        form = ExternalTopicForm()
    return render(request, "session/create_external_topic.html", {"form": form})


@login_required
def learning_view(request):
    """
    View all external learning topics.

    Returns:
        HttpResponse: List of topics.
    """
    sessions = ExternalTopic.objects.all().order_by("-created_at")
    paginator = Paginator(sessions, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, "session/learning-topic-list.html", {"sessions": page_obj})


@login_required
def edit_learning(request, learning_id):
    """
    Edit an existing learning topic.

    Args:
        learning_id (int): ID of the topic.

    Returns:
        HttpResponse: Form or redirect.
    """
    learning = get_object_or_404(ExternalTopic, id=learning_id)
    if request.method == "POST":
        form = ExternalTopicForm(request.POST, instance=learning)
        if form.is_valid():
            form.save()
            if request.user.is_staff:
                log_activity(
                    request.user,
                    f"Admin updated learning topic: '{learning.coming_soon}'.",
                    target_users=User.objects.filter(is_staff=False),
                )
            else:
                log_activity(
                    request.user, f"Updated learning topic: '{learning.coming_soon}'."
                )
            messages.success(request, "Learning updated successfully.")
            return redirect("learning-view")
    else:
        form = ExternalTopicForm(instance=learning)
    return render(
        request, "session/edit_learning.html", {"form": form, "learning": learning}
    )


@login_required
def delete_learning(request, learning_id):
    """
    Delete a learning topic.

    Args:
        learning_id (int): ID of the topic.

    Returns:
        HttpResponseRedirect: Redirect to learning view.
    """
    learning = get_object_or_404(ExternalTopic, id=learning_id)
    if request.user.is_staff:
        log_activity(
            request.user,
            f"Admin deleted learning topic: '{learning.coming_soon}'.",
            target_users=User.objects.filter(is_staff=False),
        )
    else:
        log_activity(request.user, f"Deleted learning topic: '{learning.coming_soon}'.")
    learning.delete()
    messages.success(request, "Learning deleted successfully.")
    return redirect("learning-view")


@login_required
def recent_activities(request):
    """
    View recent activities for the logged-in user.

    Returns:
        HttpResponse: List of recent activity logs.
    """
    RecentActivity.objects.filter(user=request.user, read=False).update(read=True)
    activities = RecentActivity.objects.filter(user=request.user).order_by(
        "-timestamp"
    )[:20]
    paginator = Paginator(activities, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, "session/recent_activities.html", {"activities": page_obj})


@login_required
def department_list(request):
    departments = Department.objects.all().order_by("-created_at")
    return render(request, "session/department_list.html", {"departments": departments})


@login_required
def department_create(request):
    if request.method == "POST":
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            log_activity(
                request.user,
                f"Admin created new department: '{department.name}'.",
                target_users=User.objects.filter(is_staff=False),
            )
            messages.success(request, "Department created successfully.")
            return redirect("department-list")
    else:
        form = DepartmentForm()
    return render(request, "session/department_form.html", {"form": form})


@login_required
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            log_activity(
                request.user,
                f"Admin edited department: '{department.name}'.",
                target_users=User.objects.filter(is_staff=False),
            )
            messages.success(request, "Department updated successfully.")
            return redirect("department-list")
    else:
        form = DepartmentForm(instance=department)
    return render(request, "session/department_form.html", {"form": form})


@login_required
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    department.delete()
    log_activity(
        request.user,
        f"Admin deleted department: '{department.name}'.",
        target_users=User.objects.filter(is_staff=False),
    )
    messages.success(request, "Department deleted successfully.")
    return redirect("department-list")


@staff_member_required
def export_sessions(request):
    """
    Generate and download an Excel file containing all 'Pending' session data, sorted by date.
    """
    # Fetch all pending sessions and convert to list of dicts
    sessions = SessionTopic.objects.filter(status="Pending").select_related(
        "conducted_by"
    )

    session_data = []
    for i, s in enumerate(sessions, 1):
        session_data.append(
            {
                "No.": i,
                "Date": s.date,
                "Topic": s.topic,
                "Status": s.status,
                "Assigned To": s.conducted_by.get_full_name()
                or s.conducted_by.username,
                "Place": s.place,
            }
        )

    # Create DataFrame and sort by Date
    df = pd.DataFrame(session_data)
    df.sort_values(by="Date", inplace=True)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions"

    # Write headers
    for col_num, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = column
        ws[f"{col_letter}1"].font = openpyxl.styles.Font(bold=True)

    # Write data rows
    for row_num, row in enumerate(df.itertuples(index=False), start=2):
        for col_num, value in enumerate(row, start=1):
            ws.cell(
                row=row_num,
                column=col_num,
                value=(
                    value.strftime("%Y/%m/%d")
                    if isinstance(value, pd.Timestamp)
                    else value
                ),
            )

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sessions.xlsx"'

    # Save workbook to response
    wb.save(response)
    return response


@staff_member_required
def upload_sessions_excel(request):
    """
    Handle Excel file upload to create or update SessionTopic records.
    Updates existing topics for the same user; creates new ones if no match.
    """
    if request.method == "POST":
        form = SessionUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                # Check headers
                expected_headers = [
                    "No.",
                    "Topic",
                    "Date",
                    "Status",
                    "Assigned To",
                    "Place",
                    "Cancelled Reason",
                ]
                headers = [cell.value for cell in ws[1]]
                if headers != expected_headers:
                    messages.error(
                        request,
                        "Invalid Excel file format. Expected headers: "
                        + ", ".join(expected_headers),
                    )
                    return redirect("session_list")

                # Process rows (skip header)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    (
                        no,
                        topic,
                        date_str,
                        status,
                        assigned_to,
                        place,
                        cancelled_reason,
                    ) = row

                    # Skip empty rows
                    if not topic or not assigned_to:
                        continue

                    # Parse date
                    try:
                        date = datetime.strptime(date_str, "%Y/%m/%d").date()
                    except (ValueError, TypeError):
                        messages.error(
                            request,
                            f"Invalid date format for topic '{topic}'. Expected YYYY/MM/DD.",
                        )
                        continue

                    # Find user by full name
                    try:
                        first_name, last_name = assigned_to.split(" ", 1)
                        user = User.objects.get(
                            first_name=first_name, last_name=last_name
                        )
                    except (ValueError, User.DoesNotExist):
                        messages.error(
                            request,
                            f"User '{assigned_to}' not found for topic '{topic}'.",
                        )
                        continue

                    # Validate status and place
                    valid_statuses = [choice[0] for choice in STATUSES]
                    valid_places = [choice[0] for choice in PLACE_CHOICES]
                    if status not in valid_statuses:
                        messages.error(
                            request, f"Invalid status '{status}' for topic '{topic}'."
                        )
                        continue
                    if place not in valid_places:
                        messages.error(
                            request, f"Invalid place '{place}' for topic '{topic}'."
                        )
                        continue

                    # Check for existing session by topic and user
                    session, created = SessionTopic.objects.get_or_create(
                        topic=topic,
                        conducted_by=user,
                        defaults={
                            "date": date,
                            "status": status,
                            "place": place,
                            "cancelled_reason": cancelled_reason or None,
                        },
                    )

                    if not created:
                        # Update existing session
                        session.date = date
                        session.status = status
                        session.place = place
                        session.cancelled_reason = cancelled_reason or None
                        session.save()
                        messages.info(
                            request, f"Updated session: {topic} for {assigned_to}"
                        )

                    else:
                        messages.success(
                            request, f"Created session: {topic} for {assigned_to}"
                        )

                messages.success(request, "Excel file processed successfully.")
                return redirect("session_list")

            except Exception as e:
                messages.error(request, f"Error processing Excel file: {str(e)}")
                return redirect("session_list")

        else:
            messages.error(request, "Invalid file upload.")
            return redirect("session_list")

    return redirect("session_list")
